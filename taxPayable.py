# TODO: first go with import and add the xls file in program
import openpyxl as xl
import math as m


def compute(a):
    a = float(a)
    b = 0.0
    if a < 300000:
        return b
    if 250000 < a <= 500000:
        c = a - 250000
        b += (0.05 * c)
    elif a > 500000:
        b += 12500
    if 500000 < a <= 1000000:
        c = a - 500000
        b += (0.2 * c)
    elif a > 1000000:
        b += 100000
        c = a - 1000000
        b += (0.3 * c)
    b += (0.04 * b)
    if a >= 5000000:
        b += (0.1 * b)

    return b


wb = xl.load_workbook('01export.xlsx')
sheet = wb['Sheet1']
q = sheet.cell(1, 15)
q.value = 'Tax Payable'
for i in range(2, sheet.max_row + 1):
    cell = sheet.cell(i, 14)
    tax_payable = compute(cell.value)
    taxPayable = sheet.cell(i, 15)
    taxPayable.value = m.ceil(tax_payable)
wb.save('02export.xlsx')
